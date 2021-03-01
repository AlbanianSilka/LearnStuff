import openpyxl
from openpyxl.styles import PatternFill
import sqlite3
import gdown
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import random
import timeit

gauth = GoogleAuth()
gauth.LocalWebserverAuth()
gdrive = GoogleDrive()

# ссылка на мой файл на гугл драйве
url = 'https://drive.google.com/uc?id=1Co_6XNa5fFXrm0MEi6arYfY8z6wpL7dD'
# # То, с каким именем будет выведен мой файл
output = 'Testing.xlsx'
gdown.download(url, output, quiet=False)
book = openpyxl.load_workbook('Testing.xlsx')
sheet = book.active
conn = sqlite3.connect('NewData.db')
cursor = conn.cursor()
cursor.execute('''CREATE TABLE IF NOT EXISTS result(rows_number TEXT, elapsed_time TEXT)''')
max_rows = 10  # задаём максимальное количество строк


# Функция для запуска таймера, в ней содержится декоратор
def count_time(func):
    def new_func():
        # Счёт времени делается при помощи библиотеки timeit
        time_started = timeit.default_timer()
        func()
        elapsed = timeit.default_timer() - time_started
        print('Функция "{name}" была выполнена за {time} секунд.'.format(name=func.__name__, time=round(elapsed, 10)))
        cursor.execute("INSERT INTO result(rows_number, elapsed_time) VALUES (?, ?)", (max_rows, round(elapsed, 10)))
        conn.commit()

    return new_func()


@count_time
def random_numbers():
    color = (''.join(random.choice('0123456789ABCDEF') for j in range(6))
             for i in range(1))
    color_str = ''.join(color)
    print(color_str)
    for y in range(1, max_rows):
        for x in range(1, 11):
            sheet.cell(row=y, column=x).value = random.randint(1, 100)
        # Выбираем границы закрашиваемых столбцов, границы фиксированные, поэтому чтобы продвигаться дальше -
        # задаём другие координаты
    for z in range(1, max_rows):
        x = random.randint(1, 10)
        sheet.cell(row=z, column=x).fill = PatternFill(fill_type="solid", fgColor=f"{color_str}")
    book.save('Testing.xlsx')


file1 = gdrive.CreateFile({'title': 'Testing.xlsx'})
file1.SetContentFile('Testing.xlsx')
file1.Upload()
